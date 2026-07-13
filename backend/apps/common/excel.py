"""Helpers for generating XLSX responses."""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _stringify(value):
    if value is None:
        return ''
    if isinstance(value, (datetime, date)):
        return value.strftime('%d.%m.%Y %H:%M' if isinstance(value, datetime) else '%d.%m.%Y')
    if isinstance(value, Decimal):
        return float(value)
    return value


def _get_value(row, key):
    if isinstance(row, dict):
        return row.get(key)
    return getattr(row, key, None)


def build_xlsx_response(filename, title, columns, rows, summary=None):
    """Build an XLSX response.

    columns: iterable of (key, label).
    rows: iterable of dictionaries or objects.
    summary: optional iterable of (label, value).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or 'Report'

    header_fill = PatternFill('solid', fgColor='EAF0FF')
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color='111827')

    ws.cell(row=1, column=1, value=title).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))

    start_row = 3
    for col_idx, (_, label) in enumerate(columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, (key, _) in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_stringify(_get_value(row, key)))

    last_row = ws.max_row
    if summary:
        last_row += 2
        for idx, (label, value) in enumerate(summary, start=last_row):
            ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=_stringify(value))

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.auto_filter.ref = f'A{start_row}:{get_column_letter(len(columns))}{max(ws.max_row, start_row)}'

    for col_idx, (_, label) in enumerate(columns, start=1):
        max_length = len(str(label))
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=start_row + 1, max_row=ws.max_row):
            for item in cell:
                max_length = max(max_length, len(str(item.value or '')))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 12), 42)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
