"""Views активов и склада ИС «АСУ»."""

import hashlib
from datetime import datetime
from decimal import Decimal, InvalidOperation
import unicodedata

from django.core import signing
from django.db import transaction
from django.utils.translation import gettext_lazy as _
from openpyxl import load_workbook
from rest_framework import status, viewsets, mixins
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.common.constants import (
    ASSET_TYPE_NMA,
    ASSET_TYPE_OS,
    ASSET_TYPE_TMZ,
    MOVEMENT_INVENTORY_ADJUSTMENT,
)
from apps.references.models import Asset, AssetCategory, UnitOfMeasure, Warehouse
from apps.users.access import has_access
from apps.common.permissions import IsAdmin
from apps.common.trash import SoftDeleteViewSetMixin

from .filters import AssignmentFilter, MovementFilter, WarehouseStockFilter
from .models import WarehouseStock, AssetAssignment, StockMovement, StockAlertRule, StockAlertState
from .serializers import (
    WarehouseStockSerializer,
    AssetAssignmentSerializer,
    StockMovementSerializer,
    StockAlertRuleSerializer,
    ActiveStockAlertSerializer,
)
from .services import StockAlertService, StockService


class CanViewWarehouse:
    """Просмотр складских журналов доступен только управляющему контуру."""

    def has_permission(self, request, view):
        return (
            request.user
            and request.user.is_authenticated
            and has_access(request.user, 'warehouse.view')
        )


class CanManageStockAlerts:
    """Настройки складских алармов доступны сотрудникам с правом загрузки склада."""

    def has_permission(self, request, view):
        return (
            request.user
            and request.user.is_authenticated
            and (has_access(request.user, 'warehouse.upload') or has_access(request.user, 'system.admin'))
        )

    def has_object_permission(self, request, view, obj):
        return self.has_permission(request, view)


class StockUploadView(APIView):
    """Административная загрузка остатков ТМЗ/ОС/НМА из Excel на выбранную дату."""

    permission_classes = [IsAdmin]

    # Распознаваемые заголовки столбцов (рус / каз / англ)
    COLUMN_MAP = {
        'тип': 'asset_type',
        'тип актива': 'asset_type',
        'вид актива': 'asset_type',
        'asset type': 'asset_type',
        'type': 'asset_type',
        'код': 'code',
        'код номенклатуры': 'code',
        'code': 'code',
        'артикул': 'code',
        'наименование': 'name',
        'название': 'name',
        'name': 'name',
        'наименование товара': 'name',
        'товар': 'name',
        'единицa измерения': 'unit',
        'ед. изм.': 'unit',
        'единица измерения': 'unit',
        'unit': 'unit',
        'unit of measure': 'unit',
        'количество': 'quantity',
        'кол-во': 'quantity',
        'qty': 'quantity',
        'quantity': 'quantity',
        'остаток': 'quantity',
        'цена': 'unit_price',
        'цена за единицу': 'unit_price',
        'unit price': 'unit_price',
        'price': 'unit_price',
        'сумма': 'total_amount',
        'total': 'total_amount',
        'total amount': 'total_amount',
        'сумма всего': 'total_amount',
        'категория': 'category',
        'category': 'category',
        'группа': 'category',
        'место хранения': 'location',
        'location': 'location',
        'склад': 'location',
    }

    ASSET_TYPE_MAP = {
        'тмз': ASSET_TYPE_TMZ,
        'tmz': ASSET_TYPE_TMZ,
        'ос': ASSET_TYPE_OS,
        'os': ASSET_TYPE_OS,
        'нма': ASSET_TYPE_NMA,
        'nma': ASSET_TYPE_NMA,
    }

    ASSET_CODE_PREFIXES = {
        ASSET_TYPE_TMZ: 'ТМЗ',
        ASSET_TYPE_OS: 'ОС',
        ASSET_TYPE_NMA: 'НМА',
    }

    ASSET_CODE_PREFIX_ALIASES = {
        ASSET_TYPE_TMZ: ('ТМЗ', 'TMZ'),
        ASSET_TYPE_OS: ('ОС', 'OS'),
        ASSET_TYPE_NMA: ('НМА', 'NMA'),
    }

    PREVIEW_TOKEN_SALT = 'assets.stock-upload.preview'
    PREVIEW_TOKEN_MAX_AGE = 60 * 60

    def post(self, request):
        """Первый этап: проверить файл и вернуть предпросмотр без записи в БД."""
        return self.preview(request)

    def preview(self, request):
        try:
            upload = self._load_upload(request)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rows, summary, errors = self._parse_upload_rows(
                upload['worksheet'],
                upload['default_warehouse'],
            )
        finally:
            upload['workbook'].close()

        token = signing.dumps(
            self._preview_payload(upload, request.user.pk),
            salt=self.PREVIEW_TOKEN_SALT,
            compress=True,
        )
        return Response({
            'success': True,
            'stage': 'preview',
            'file_name': upload['file_name'],
            'balance_date': upload['balance_date_str'] or None,
            'preview_token': token,
            'can_confirm': not errors and bool(rows),
            'summary': summary,
            'rows': self._serialize_preview_rows(rows),
            'errors': errors,
        })

    def confirm(self, request):
        """Второй этап: повторно проверить файл и атомарно записать его в БД."""
        preview_token = request.data.get('preview_token')
        if not preview_token:
            return Response(
                {'detail': _('Сначала выполните проверку файла')},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            upload = self._load_upload(request)
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            signed_payload = signing.loads(
                preview_token,
                salt=self.PREVIEW_TOKEN_SALT,
                max_age=self.PREVIEW_TOKEN_MAX_AGE,
            )
        except signing.SignatureExpired:
            upload['workbook'].close()
            return Response(
                {'detail': _('Срок действия проверки истек. Проверьте файл повторно')},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except signing.BadSignature:
            upload['workbook'].close()
            return Response(
                {'detail': _('Некорректное подтверждение загрузки')},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if signed_payload != self._preview_payload(upload, request.user.pk):
            upload['workbook'].close()
            return Response(
                {
                    'detail': _(
                        'Файл, дата остатка или склад изменились после проверки. '
                        'Выполните проверку повторно'
                    ),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            rows, _summary, errors = self._parse_upload_rows(
                upload['worksheet'],
                upload['default_warehouse'],
            )
        finally:
            upload['workbook'].close()

        if errors:
            return Response(
                {
                    'detail': _('Файл содержит ошибки и не может быть загружен'),
                    'errors': errors,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        result = self._commit_upload_rows(
            rows=rows,
            balance_date=upload['balance_date'],
            default_warehouse=upload['default_warehouse'],
            user=request.user,
        )
        return Response(result)

    def _load_upload(self, request):
        balance_date_str = request.query_params.get('balance_date') or request.data.get('balance_date')
        balance_date = None
        if balance_date_str:
            try:
                balance_date = datetime.strptime(balance_date_str, '%Y-%m-%d').date()
            except ValueError:
                raise ValueError(_('balance_date должен быть в формате YYYY-MM-DD'))

        warehouse_id = request.query_params.get('warehouse') or request.data.get('warehouse')
        default_warehouse = None
        if warehouse_id:
            default_warehouse = Warehouse.objects.filter(pk=warehouse_id, is_active=True).first()
            if not default_warehouse:
                raise ValueError(_('Склад не найден или не активен'))

        file_obj = request.FILES.get('file')
        if not file_obj:
            raise ValueError(_('Необходимо приложить файл'))

        file_digest = self._file_digest(file_obj)

        try:
            file_obj.seek(0)
            wb = load_workbook(file_obj, data_only=True)
            ws = wb.active
        except Exception as exc:
            raise ValueError(str(_('Не удалось прочитать Excel: ')) + str(exc))

        if not ws or ws.max_row < 2:
            wb.close()
            raise ValueError(_('Файл пуст или не содержит данных'))

        headers = self._parse_headers(ws[1])
        required = {'asset_type', 'code', 'name', 'unit', 'quantity'}
        missing = required - set(headers.values())
        if missing:
            column_names = {
                'asset_type': _('Тип актива'),
                'code': _('Код номенклатуры'),
                'name': _('Наименование'),
                'unit': _('Единица измерения'),
                'quantity': _('Количество'),
            }
            wb.close()
            raise ValueError(
                str(_('Отсутствуют обязательные столбцы: '))
                + ', '.join(str(column_names[field]) for field in sorted(missing))
            )

        return {
            'balance_date': balance_date,
            'balance_date_str': balance_date.isoformat() if balance_date else '',
            'default_warehouse': default_warehouse,
            'warehouse_id': default_warehouse.pk if default_warehouse else None,
            'file_digest': file_digest,
            'file_name': file_obj.name,
            'workbook': wb,
            'worksheet': ws,
        }

    @staticmethod
    def _file_digest(file_obj):
        digest = hashlib.sha256()
        file_obj.seek(0)
        for chunk in file_obj.chunks():
            digest.update(chunk)
        file_obj.seek(0)
        return digest.hexdigest()

    @staticmethod
    def _preview_payload(upload, user_id):
        return {
            'file_digest': upload['file_digest'],
            'balance_date': upload['balance_date_str'],
            'warehouse_id': upload['warehouse_id'],
            'user_id': user_id,
        }

    def _parse_upload_rows(self, worksheet, default_warehouse):
        headers = self._parse_headers(worksheet[1])
        self._initialize_reference_resolution()
        rows = []
        for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value is not None and str(value).strip() for value in row):
                continue

            values = dict(zip(headers.keys(), row))
            mapped = {headers[col]: values[col] for col in headers if values[col] is not None}
            row_errors = []
            row_warnings = []

            asset_type = self._resolve_asset_type(mapped.get('asset_type'))
            source_code = self._clean_name(mapped.get('code'))
            code = self._normalize_asset_code(source_code, asset_type) if asset_type else source_code
            name = self._clean_name(mapped.get('name'))
            unit = self._clean_name(mapped.get('unit'))

            if not asset_type:
                row_errors.append('Некорректный тип актива. Допустимые значения: ОС, НМА, ТМЗ')
            if not source_code:
                row_errors.append('Не указан код номенклатуры')
            if not name:
                row_errors.append('Не указано наименование')
            if not unit:
                row_errors.append('Не указана единица измерения')
            if code and len(code) > Asset._meta.get_field('code').max_length:
                row_errors.append('Код номенклатуры превышает допустимую длину')
            if len(name) > Asset._meta.get_field('name').max_length:
                row_errors.append('Наименование превышает допустимую длину')
            if len(unit) > Asset._meta.get_field('unit_of_measure').max_length:
                row_errors.append('Единица измерения превышает допустимую длину')

            quantity = None
            try:
                quantity = self._to_decimal(mapped.get('quantity'), 2)
                if quantity is None or quantity < 0 or quantity >= Decimal('10000000000'):
                    raise ValueError
            except (ValueError, InvalidOperation):
                row_errors.append('Некорректное количество')

            unit_price = None
            total_amount = None
            try:
                unit_price = self._to_decimal(mapped.get('unit_price'), 2)
                total_amount = self._to_decimal(mapped.get('total_amount'), 2)
                if (
                    (unit_price is not None and (unit_price < 0 or unit_price >= Decimal('10000000000000')))
                    or (
                        total_amount is not None
                        and (total_amount < 0 or total_amount >= Decimal('10000000000000'))
                    )
                ):
                    raise ValueError
            except (ValueError, InvalidOperation):
                row_errors.append('Некорректная цена или сумма')

            supplied_price = unit_price
            supplied_total = total_amount
            if quantity is not None and 'Некорректная цена или сумма' not in row_errors:
                if total_amount is None and unit_price is not None:
                    total_amount = (unit_price * quantity).quantize(Decimal('0.01'))
                elif unit_price is None and total_amount is not None and quantity:
                    unit_price = (total_amount / quantity).quantize(Decimal('0.01'))
                elif unit_price is None:
                    unit_price = Decimal('0')
                if total_amount is None:
                    total_amount = (unit_price * quantity).quantize(Decimal('0.01'))
                if supplied_price is not None and supplied_total is not None:
                    difference = abs(supplied_total - (quantity * supplied_price))
                    if difference > Decimal('0.01'):
                        row_warnings.append(
                            f'Сумма отличается от количества × цены на {difference.quantize(Decimal("0.01"))}'
                        )

            category = self._clean_name(mapped.get('category'))
            if asset_type and not category:
                category = f'Загружено из Excel ({asset_type})'
            location_input = self._clean_name(mapped.get('location'))
            if len(category) > AssetCategory._meta.get_field('name').max_length:
                row_errors.append('Категория превышает допустимую длину')
            if len(location_input) > Warehouse._meta.get_field('name').max_length:
                row_errors.append('Место хранения превышает допустимую длину')

            existing_warehouse = (
                self._warehouse_cache.get(self._name_key(location_input))
                if location_input
                else default_warehouse
            )
            warehouse_name = (
                existing_warehouse.name
                if existing_warehouse
                else location_input
            )
            rows.append({
                'excel_row': idx,
                'asset_type': asset_type or self._clean_name(mapped.get('asset_type')),
                'source_code': source_code,
                'code': code,
                'name': name,
                'unit': unit,
                'quantity': str(quantity) if quantity is not None else self._clean_name(mapped.get('quantity')),
                'unit_price': str(unit_price) if unit_price is not None else '',
                'total_amount': str(total_amount) if total_amount is not None else '',
                'category': category,
                'warehouse': warehouse_name,
                'location_input': location_input,
                'errors': row_errors,
                'warnings': row_warnings,
                'new_references': [],
                'action': '',
                'status': '',
                '_quantity': quantity,
                '_unit_price': unit_price,
                '_total_amount': total_amount,
            })

        code_rows = {}
        for item in rows:
            if item['code'] and not item['errors']:
                code_rows.setdefault(item['code'], []).append(item)
        for code, duplicate_rows in code_rows.items():
            if len(duplicate_rows) > 1:
                row_numbers = ', '.join(str(item['excel_row']) for item in duplicate_rows)
                for item in duplicate_rows:
                    item['errors'].append(
                        f'Код {code} повторяется в строках {row_numbers}'
                    )

        valid_codes = [item['code'] for item in rows if not item['errors']]
        existing_asset_codes = set(
            Asset.objects.filter(code__in=valid_codes).values_list('code', flat=True)
        )
        new_units = {}
        new_warehouses = {}
        new_categories = {}
        type_counts = {}

        for item in rows:
            if item['errors']:
                item['status'] = 'error'
                continue

            item['action'] = 'update' if item['code'] in existing_asset_codes else 'create'
            type_counts[item['asset_type']] = type_counts.get(item['asset_type'], 0) + 1

            unit_key = self._name_key(item['unit'])
            if unit_key not in self._unit_cache:
                new_units.setdefault(unit_key, {'name': item['unit']})
                item['new_references'].append(f'Единица измерения: {item["unit"]}')

            category_key = (item['asset_type'], self._name_key(item['category']))
            if category_key not in self._category_cache:
                new_categories.setdefault(category_key, {
                    'name': item['category'],
                    'asset_type': item['asset_type'],
                })
                item['new_references'].append(f'Категория: {item["category"]}')

            if item['location_input']:
                warehouse_key = self._name_key(item['location_input'])
                if warehouse_key not in self._warehouse_cache:
                    new_warehouses.setdefault(warehouse_key, {'name': item['location_input']})
                    item['new_references'].append(f'Склад: {item["location_input"]}')

            item['status'] = 'new_references' if item['new_references'] else 'ready'

        errors = [
            {'row': item['excel_row'], 'detail': '; '.join(item['errors'])}
            for item in rows
            if item['errors']
        ]
        valid_rows = [item for item in rows if not item['errors']]
        summary = {
            'total_rows': len(rows),
            'valid_rows': len(valid_rows),
            'invalid_rows': len(errors),
            'warning_rows': sum(bool(item['warnings']) for item in valid_rows),
            'create_assets': sum(item['action'] == 'create' for item in valid_rows),
            'update_assets': sum(item['action'] == 'update' for item in valid_rows),
            'asset_types': type_counts,
            'new_references': {
                'units': list(new_units.values()),
                'warehouses': list(new_warehouses.values()),
                'categories': list(new_categories.values()),
            },
        }
        return rows, summary, errors

    @staticmethod
    def _serialize_preview_rows(rows):
        return [
            {key: value for key, value in item.items() if not key.startswith('_')}
            for item in rows
        ]

    def _commit_upload_rows(self, rows, balance_date, default_warehouse, user):
        self._initialize_reference_resolution()
        created_assets = 0
        updated_assets = 0
        created_stock = 0
        updated_stock = 0
        processed_asset_types = []

        with transaction.atomic():
            for item in rows:
                unit_ref = self._resolve_unit(item['unit'])
                category = self._resolve_category(item['category'], item['asset_type'])
                warehouse = (
                    self._resolve_warehouse(item['location_input'])
                    if item['location_input']
                    else default_warehouse
                )
                location = warehouse.name if warehouse else ''

                asset, asset_created = Asset.objects.update_or_create(
                    code=item['code'],
                    defaults={
                        'name': item['name'],
                        'asset_type': item['asset_type'],
                        'category': category,
                        'unit_of_measure': unit_ref.name,
                        'unit_of_measure_ref': unit_ref,
                        'unit_price': item['_unit_price'],
                        'balance_date': balance_date,
                    },
                )
                if asset_created:
                    created_assets += 1
                else:
                    updated_assets += 1

                existing_stock = WarehouseStock.objects.select_for_update().filter(
                    asset=asset,
                    warehouse=warehouse,
                ).first()
                old_quantity = existing_stock.quantity if existing_stock else Decimal('0')
                old_total = existing_stock.total_amount if existing_stock else Decimal('0')

                stock, stock_created = WarehouseStock.objects.update_or_create(
                    asset=asset,
                    warehouse=warehouse,
                    defaults={
                        'quantity': item['_quantity'],
                        'total_amount': item['_total_amount'],
                        'unit_price': item['_unit_price'],
                        'balance_date': balance_date,
                        'location': location or (existing_stock.location if existing_stock else ''),
                    },
                )
                if stock_created:
                    created_stock += 1
                else:
                    updated_stock += 1

                quantity_delta = stock.quantity - old_quantity
                total_delta = stock.total_amount - old_total
                if quantity_delta or stock_created:
                    StockMovement.objects.create(
                        asset=asset,
                        movement_type=MOVEMENT_INVENTORY_ADJUSTMENT,
                        quantity=quantity_delta,
                        unit_price=item['_unit_price'],
                        total_amount=total_delta,
                        performed_by=user,
                        warehouse=stock.warehouse,
                        comment=_('Корректировка остатка по загрузке Excel'),
                    )
                if item['asset_type'] not in processed_asset_types:
                    processed_asset_types.append(item['asset_type'])

        return {
            'success': True,
            'stage': 'confirmed',
            'asset_type': processed_asset_types[0] if len(processed_asset_types) == 1 else None,
            'asset_types': processed_asset_types,
            'balance_date': balance_date,
            'processed': len(rows),
            'skipped': 0,
            'created_assets': created_assets,
            'updated_assets': updated_assets,
            'created_stock': created_stock,
            'updated_stock': updated_stock,
            'created_references': self._created_references,
            'errors': [],
        }

    def _parse_headers(self, header_row):
        headers = {}
        for idx, cell in enumerate(header_row):
            if cell.value is None:
                continue
            key = ' '.join(str(cell.value).strip().casefold().split())
            mapped = self.COLUMN_MAP.get(key)
            if mapped:
                headers[idx] = mapped
        return headers

    def _to_decimal(self, value, places):
        if value is None or value == '':
            return None
        if isinstance(value, (int, float)):
            return Decimal(str(value)).quantize(Decimal('0.1') ** places)
        return Decimal(str(value).replace(' ', '').replace(',', '.')).quantize(Decimal('0.1') ** places)

    @staticmethod
    def _clean_name(value):
        """Читабельное значение: Unicode NFKC и один пробел между словами."""
        if value is None:
            return ''
        return ' '.join(unicodedata.normalize('NFKC', str(value)).split())

    @classmethod
    def _name_key(cls, value):
        """Ключ точного сопоставления без учета регистра и любых пробелов."""
        return ''.join(cls._clean_name(value).split()).casefold()

    @classmethod
    def _resolve_asset_type(cls, value):
        return cls.ASSET_TYPE_MAP.get(cls._name_key(value))

    @classmethod
    def _normalize_asset_code(cls, value, asset_type):
        """Добавить обязательный кириллический префикс типа без дублирования."""
        code = cls._clean_name(value)
        if not code:
            return ''

        for alias in cls.ASSET_CODE_PREFIX_ALIASES.get(asset_type, ()):
            if code.casefold().startswith(alias.casefold()):
                code = code[len(alias):].lstrip(' -_')
                break

        if not code:
            return ''
        return f'{cls.ASSET_CODE_PREFIXES[asset_type]}{code}'

    def _initialize_reference_resolution(self):
        self._unit_cache = {}
        for unit in UnitOfMeasure.objects.order_by('id'):
            self._unit_cache.setdefault(self._name_key(unit.name), unit)

        self._warehouse_cache = {}
        for warehouse in Warehouse.objects.order_by('id'):
            self._warehouse_cache.setdefault(self._name_key(warehouse.name), warehouse)

        self._category_cache = {}
        for category in AssetCategory.objects.order_by('id'):
            key = (category.asset_type, self._name_key(category.name))
            self._category_cache.setdefault(key, category)

        self._created_references = {
            'units': [],
            'warehouses': [],
            'categories': [],
        }

    def _ensure_reference_resolution(self):
        if not hasattr(self, '_created_references'):
            self._initialize_reference_resolution()

    @staticmethod
    def _next_reference_code(model, prefix):
        sequence = model.objects.count() + 1
        while True:
            code = f'{prefix}-{sequence:04d}'
            if not model.objects.filter(code=code).exists():
                return code
            sequence += 1

    def _resolve_unit(self, name):
        self._ensure_reference_resolution()
        display_name = self._clean_name(name)
        key = self._name_key(display_name)
        if not key:
            return None
        unit = self._unit_cache.get(key)
        if unit:
            return unit
        unit = UnitOfMeasure.objects.create(
            name=display_name,
            code=self._next_reference_code(UnitOfMeasure, 'UOM'),
        )
        self._unit_cache[key] = unit
        self._created_references['units'].append({
            'id': unit.pk,
            'name': unit.name,
            'code': unit.code,
        })
        return unit

    def _resolve_warehouse(self, name):
        self._ensure_reference_resolution()
        display_name = self._clean_name(name)
        key = self._name_key(display_name)
        if not key:
            return None
        warehouse = self._warehouse_cache.get(key)
        if warehouse:
            return warehouse
        warehouse = Warehouse.objects.create(
            name=display_name,
            code=self._next_reference_code(Warehouse, 'WH'),
        )
        self._warehouse_cache[key] = warehouse
        self._created_references['warehouses'].append({
            'id': warehouse.pk,
            'name': warehouse.name,
            'code': warehouse.code,
        })
        return warehouse

    def _resolve_category(self, name, asset_type):
        self._ensure_reference_resolution()
        display_name = self._clean_name(name) or f'Загружено из Excel ({asset_type})'
        key = (asset_type, self._name_key(display_name))
        category = self._category_cache.get(key)
        if category:
            return category
        category = AssetCategory.objects.create(
            name=display_name,
            asset_type=asset_type,
            code=self._next_reference_code(AssetCategory, 'CAT'),
        )
        self._category_cache[key] = category
        self._created_references['categories'].append({
            'id': category.pk,
            'name': category.name,
            'code': category.code,
            'asset_type': category.asset_type,
        })
        return category


class StockUploadConfirmView(StockUploadView):
    """Подтверждение ранее проверенной загрузки остатков."""

    def post(self, request):
        return self.confirm(request)


class WarehouseStockViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin,
                            viewsets.GenericViewSet):
    """Просмотр остатков на складе."""
    queryset = WarehouseStock.objects.select_related('asset', 'asset__category', 'asset__group', 'warehouse').all()
    serializer_class = WarehouseStockSerializer
    permission_classes = [CanViewWarehouse]
    filterset_class = WarehouseStockFilter
    search_fields = ['asset__name', 'asset__code', 'location', 'warehouse__name']
    ordering_fields = ['quantity', 'total_amount', 'updated_at']


class AssetAssignmentViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin,
                             mixins.CreateModelMixin, viewsets.GenericViewSet):
    """Просмотр и создание закреплений активов."""
    queryset = AssetAssignment.objects.select_related(
        'asset', 'asset__category', 'asset__group', 'user', 'assigned_by', 'warehouse',
    ).all()
    serializer_class = AssetAssignmentSerializer
    permission_classes = [CanViewWarehouse]
    filterset_class = AssignmentFilter
    search_fields = [
        'asset__name', 'asset__code', 'asset__inventory_number',
        'user__username', 'user__last_name', 'user__first_name', 'user__patronymic',
    ]
    ordering_fields = ['assigned_at', 'status']

    def perform_create(self, serializer):
        serializer.save(assigned_by=self.request.user)

    @action(detail=True, methods=['post'], url_path='release')
    def release(self, request, pk=None):
        """Административно снять актив с сотрудника без удаления истории."""
        if not has_access(request.user, 'system.admin'):
            return Response(
                {'detail': _('Снимать закрепления может только администратор')},
                status=status.HTTP_403_FORBIDDEN,
            )
        try:
            assignment = StockService.release_assignment(
                self.get_object(),
                released_by=request.user,
                reason=request.data.get('reason', ''),
            )
        except ValueError as exc:
            return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(AssetAssignmentSerializer(assignment).data)


class StockMovementViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin,
                           viewsets.GenericViewSet):
    """Просмотр журнала движения активов."""
    queryset = StockMovement.objects.select_related(
        'asset', 'asset__category', 'asset__group', 'from_user', 'to_user', 'performed_by', 'warehouse',
    ).all()
    serializer_class = StockMovementSerializer
    permission_classes = [CanViewWarehouse]
    filterset_class = MovementFilter
    search_fields = ['asset__name', 'comment']
    ordering_fields = ['performed_at', 'total_amount']


class StockAlertRuleViewSet(SoftDeleteViewSetMixin, viewsets.ModelViewSet):
    """Настройки критических остатков."""

    queryset = StockAlertRule.objects.prefetch_related(
        'recipients', 'groups', 'assets', 'warehouses', 'states',
    ).all()
    serializer_class = StockAlertRuleSerializer
    permission_classes = [CanManageStockAlerts]
    search_fields = ['name', 'message_template', 'assets__name', 'groups__name']
    ordering_fields = ['name', 'threshold_quantity', 'updated_at']
    ordering = ['name']

    def perform_create(self, serializer):
        rule = serializer.save()
        StockAlertService.evaluate_rule(rule)

    def perform_update(self, serializer):
        rule = serializer.save()
        StockAlertService.evaluate_rule(rule)


class ActiveStockAlertView(APIView):
    """Активные складские алармы текущего пользователя."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        alerts = StockAlertState.objects.select_related(
            'rule', 'stock', 'stock__asset', 'stock__warehouse',
        ).filter(
            is_active=True,
            rule__is_active=True,
            rule__recipients=request.user,
        ).order_by('stock__asset__name').distinct()
        return Response(ActiveStockAlertSerializer(alerts, many=True).data)
