import secrets
import string
from collections import Counter
from io import BytesIO

from django.contrib.auth import password_validation
from django.core.exceptions import ValidationError as DjangoValidationError
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import generics, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import AllowAny, BasePermission, IsAuthenticated, SAFE_METHODS
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken

from apps.common.constants import ROLE_CHOICES

from .filters import UserFilter
from .access import ACCESS_DEFINITIONS, ROLE_DEFAULT_ACCESS, effective_access_detail, has_access, normalize_position
from .models import Department, PositionAccessRule, User, UserAccessOverride
from .serializers import (
    AccessDefinitionSerializer,
    DepartmentSerializer,
    DepartmentTreeSerializer,
    LoginSerializer,
    PositionAccessRuleSerializer,
    UserAccessOverrideSerializer,
    UserCreateSerializer,
    UserProfileSerializer,
    UserSerializer,
)


class CanManageUsers(BasePermission):
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        action = getattr(view, 'action', '')
        if action in ('list', 'retrieve') and has_access(request.user, 'access.manage'):
            return True
        return has_access(request.user, 'users.manage')


class CanManageAccess(BasePermission):
    def has_permission(self, request, view):
        return bool(request.user and request.user.is_authenticated and has_access(request.user, 'access.manage'))


ACCESS_MATRIX_SOURCE_LABELS = {
    'none': 'Нет правила',
    'role': 'По роли',
    'manager': 'По статусу управляющего',
    'position_allow': 'Разрешено по должности',
    'position_deny': 'Запрещено по должности',
    'user_grant': 'Разрешено индивидуально',
    'user_deny': 'Запрещено индивидуально',
}


def access_state_label(value):
    if value is True:
        return 'Активно'
    if value is False:
        return 'Запрещено'
    return 'Не указано'


def effective_access_state_label(permission):
    if permission.get('allowed'):
        return 'Активно'
    if permission.get('source') in ('position_deny', 'user_deny'):
        return 'Запрещено'
    return 'Не указано'


def user_display_name(user):
    return user.get_full_name() or user.username


def write_access_sheet(ws, title, columns, rows):
    header_fill = PatternFill('solid', fgColor='EAF0FF')
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color='111827')

    ws.title = title[:31] or 'Sheet'
    ws.cell(row=1, column=1, value=title).font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))

    header_row = 3
    for col_idx, (_, label) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row_idx, row in enumerate(rows, start=header_row + 1):
        for col_idx, (key, _) in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key, ''))
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    if columns:
        ws.auto_filter.ref = f'A{header_row}:{get_column_letter(len(columns))}{max(ws.max_row, header_row)}'

    for col_idx, (_, label) in enumerate(columns, start=1):
        max_length = len(str(label))
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=header_row + 1, max_row=ws.max_row):
            for item in cell:
                max_length = max(max_length, len(str(item.value or '')))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 12), 42)


class ReadOnlyOrCanManageDepartments(BasePermission):
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        if request.method in SAFE_METHODS:
            return True
        return has_access(request.user, 'users.manage') or has_access(request.user, 'references.manage')


class LoginView(APIView):
    """Authenticate user and issue JWT tokens."""

    permission_classes = [AllowAny]

    def post(self, request):
        serializer = LoginSerializer(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)

        user = serializer.validated_data['user']
        refresh = RefreshToken.for_user(user)

        return Response(
            {
                'access': str(refresh.access_token),
                'refresh': str(refresh),
                'user': UserProfileSerializer(user, context={'request': request}).data,
            },
            status=status.HTTP_200_OK,
        )


class LogoutView(APIView):
    """Invalidate refresh token."""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            refresh_token = request.data.get('refresh')
            if refresh_token:
                token = RefreshToken(refresh_token)
                token.blacklist()
            return Response(status=status.HTTP_205_RESET_CONTENT)
        except Exception:
            return Response(status=status.HTTP_400_BAD_REQUEST)


class CurrentUserView(generics.RetrieveUpdateAPIView):
    """Retrieve and update the current user's profile."""

    serializer_class = UserProfileSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get_object(self):
        return self.request.user

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context


class UserViewSet(viewsets.ModelViewSet):
    """User CRUD for admins."""

    queryset = User.objects.select_related('department', 'position_ref', 'supervisor').all()
    permission_classes = [CanManageUsers]
    filterset_class = UserFilter
    search_fields = [
        'username', 'first_name', 'last_name', 'email', 'patronymic',
        'phone', 'position', 'department__name', 'supervisor__last_name',
    ]
    ordering_fields = ['last_name', 'date_joined', 'last_login', 'role', 'is_active']
    ordering = ['last_name', 'first_name']
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        return UserSerializer

    @action(detail=True, methods=['post'], url_path='reset-password')
    def reset_password(self, request, pk=None):
        user = self.get_object()
        alphabet = string.ascii_letters + string.digits + '!@#$%'
        temp_password = ''.join(secrets.choice(alphabet) for _ in range(12))
        user.set_password(temp_password)
        user.save(update_fields=['password'])
        return Response(
            {
                'detail': f'Временный пароль: {temp_password}',
                'temp_password': temp_password,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=['post'], url_path='change-password')
    def change_password(self, request, pk=None):
        user = self.get_object()
        password = request.data.get('password') or request.data.get('new_password')

        if not password:
            return Response(
                {'password': ['Введите новый пароль.']},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            password_validation.validate_password(password, user)
        except DjangoValidationError as exc:
            return Response({'password': list(exc.messages)}, status=status.HTTP_400_BAD_REQUEST)

        user.set_password(password)
        user.save(update_fields=['password'])
        return Response({'detail': 'Пароль пользователя изменен.'}, status=status.HTTP_200_OK)


class AccessDefinitionView(APIView):
    """Available permissions and role defaults."""

    permission_classes = [CanManageAccess]

    def get(self, request):
        return Response({
            'permissions': AccessDefinitionSerializer(ACCESS_DEFINITIONS, many=True).data,
            'role_defaults': {
                role: sorted(codes)
                for role, codes in ROLE_DEFAULT_ACCESS.items()
            },
        })


class EffectiveUserAccessView(APIView):
    """Effective permissions for one user."""

    permission_classes = [CanManageAccess]

    def get(self, request, user_id):
        user = User.objects.select_related('department', 'position_ref', 'supervisor').get(pk=user_id)
        return Response(effective_access_detail(user))


class AccessMatrixExportView(APIView):
    """Download complete access matrix as XLSX."""

    permission_classes = [CanManageAccess]

    def get(self, request):
        role_labels = {role: str(label) for role, label in ROLE_CHOICES}
        role_codes = [role for role, _ in ROLE_CHOICES if role in ROLE_DEFAULT_ACCESS]
        definitions = list(ACCESS_DEFINITIONS)
        users = list(
            User.objects.select_related('department', 'position_ref', 'supervisor')
            .order_by('last_name', 'first_name', 'username')
        )
        position_rules = list(
            PositionAccessRule.objects
            .order_by('position', 'permission_code', '-updated_at')
        )
        user_overrides = list(
            UserAccessOverride.objects.select_related('user', 'user__department', 'user__position_ref')
            .order_by('user__last_name', 'user__first_name', 'permission_code')
        )
        definition_by_code = {item.code: item for item in definitions}

        position_names = {}
        for rule in position_rules:
            normalized = normalize_position(rule.position)
            if normalized and normalized not in position_names:
                position_names[normalized] = rule.position
        for user in users:
            normalized = normalize_position(user.position)
            if normalized and normalized not in position_names:
                position_names[normalized] = user.position
        positions = sorted(position_names.items(), key=lambda item: item[1].casefold())
        employee_counts = Counter(normalize_position(user.position) for user in users if normalize_position(user.position))
        active_position_rules = {
            (normalize_position(rule.position), rule.permission_code): rule
            for rule in position_rules
            if rule.is_active
        }

        wb = Workbook()

        role_columns = [
            ('category', 'Раздел'),
            ('permission_code', 'Код права'),
            ('permission_name', 'Право'),
            ('description', 'Описание'),
        ] + [(f'role_{role}', role_labels.get(role, role)) for role in role_codes]
        role_rows = []
        for definition in definitions:
            row = {
                'category': definition.category,
                'permission_code': definition.code,
                'permission_name': definition.name,
                'description': definition.description,
            }
            for role in role_codes:
                row[f'role_{role}'] = access_state_label(definition.code in ROLE_DEFAULT_ACCESS.get(role, set()))
            role_rows.append(row)
        write_access_sheet(wb.active, 'Права ролей', role_columns, role_rows)

        position_matrix_columns = [
            ('category', 'Раздел'),
            ('permission_code', 'Код права'),
            ('permission_name', 'Право'),
        ] + [(f'position_{idx}', name) for idx, (_, name) in enumerate(positions)]
        position_matrix_rows = []
        for definition in definitions:
            row = {
                'category': definition.category,
                'permission_code': definition.code,
                'permission_name': definition.name,
            }
            for idx, (normalized, _) in enumerate(positions):
                rule = active_position_rules.get((normalized, definition.code))
                row[f'position_{idx}'] = access_state_label(None if not rule else rule.is_allowed)
            position_matrix_rows.append(row)
        write_access_sheet(wb.create_sheet(), 'Должности матрица', position_matrix_columns, position_matrix_rows)

        position_detail_columns = [
            ('position', 'Должность'),
            ('employee_count', 'Сотрудников'),
            ('category', 'Раздел'),
            ('permission_code', 'Код права'),
            ('permission_name', 'Право'),
            ('state', 'Состояние'),
            ('comment', 'Комментарий'),
            ('updated_at', 'Обновлено'),
        ]
        position_detail_rows = []
        for normalized, position_name in positions:
            for definition in definitions:
                rule = active_position_rules.get((normalized, definition.code))
                position_detail_rows.append({
                    'position': position_name,
                    'employee_count': employee_counts.get(normalized, 0),
                    'category': definition.category,
                    'permission_code': definition.code,
                    'permission_name': definition.name,
                    'state': access_state_label(None if not rule else rule.is_allowed),
                    'comment': rule.comment if rule else '',
                    'updated_at': timezone.localtime(rule.updated_at).strftime('%d.%m.%Y %H:%M') if rule else '',
                })
        write_access_sheet(wb.create_sheet(), 'Должности детали', position_detail_columns, position_detail_rows)

        override_columns = [
            ('user_name', 'Сотрудник'),
            ('username', 'Логин'),
            ('department', 'Подразделение'),
            ('position', 'Должность'),
            ('role', 'Роль'),
            ('category', 'Раздел'),
            ('permission_code', 'Код права'),
            ('permission_name', 'Право'),
            ('state', 'Индивидуальное правило'),
            ('comment', 'Комментарий'),
            ('updated_at', 'Обновлено'),
        ]
        override_rows = []
        for override in user_overrides:
            definition = definition_by_code.get(override.permission_code)
            override_rows.append({
                'user_name': user_display_name(override.user),
                'username': override.user.username,
                'department': override.user.department.name if override.user.department else '',
                'position': override.user.position,
                'role': override.user.get_role_display(),
                'category': definition.category if definition else '',
                'permission_code': override.permission_code,
                'permission_name': definition.name if definition else override.permission_code,
                'state': 'Активно' if override.mode == UserAccessOverride.MODE_GRANT else 'Запрещено',
                'comment': override.comment,
                'updated_at': timezone.localtime(override.updated_at).strftime('%d.%m.%Y %H:%M'),
            })
        write_access_sheet(wb.create_sheet(), 'Индивидуальные права', override_columns, override_rows)

        effective_columns = [
            ('user_name', 'Сотрудник'),
            ('username', 'Логин'),
            ('department', 'Подразделение'),
            ('position', 'Должность'),
            ('role', 'Роль'),
            ('category', 'Раздел'),
            ('permission_code', 'Код права'),
            ('permission_name', 'Право'),
            ('effective_state', 'Итоговый доступ'),
            ('source', 'Источник'),
        ]
        effective_rows = []
        for user in users:
            detail = effective_access_detail(user)
            for permission in detail['permissions']:
                effective_rows.append({
                    'user_name': user_display_name(user),
                    'username': user.username,
                    'department': user.department.name if user.department else '',
                    'position': detail['position'],
                    'role': user.get_role_display(),
                    'category': permission['category'],
                    'permission_code': permission['code'],
                    'permission_name': permission['name'],
                    'effective_state': effective_access_state_label(permission),
                    'source': ACCESS_MATRIX_SOURCE_LABELS.get(permission['source'], permission['source']),
                })
        write_access_sheet(wb.create_sheet(), 'Итог сотрудников', effective_columns, effective_rows)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="access_matrix_{timezone.localdate().isoformat()}.xlsx"'
        return response


class PositionAccessRuleViewSet(viewsets.ModelViewSet):
    """CRUD for permissions inherited by position."""

    queryset = PositionAccessRule.objects.all()
    serializer_class = PositionAccessRuleSerializer
    permission_classes = [CanManageAccess]
    search_fields = ['position', 'permission_code', 'comment']
    ordering_fields = ['position', 'permission_code', 'is_allowed', 'is_active']
    ordering = ['position', 'permission_code']


class UserAccessOverrideViewSet(viewsets.ModelViewSet):
    """CRUD for personal permission overrides."""

    queryset = UserAccessOverride.objects.select_related('user').all()
    serializer_class = UserAccessOverrideSerializer
    permission_classes = [CanManageAccess]
    filterset_fields = ['user', 'permission_code', 'mode']
    search_fields = ['user__username', 'user__last_name', 'user__first_name', 'permission_code', 'comment']
    ordering_fields = ['user__last_name', 'permission_code', 'mode']
    ordering = ['user__last_name', 'user__first_name', 'permission_code']


class DepartmentViewSet(viewsets.ModelViewSet):
    """Department CRUD."""

    queryset = Department.objects.select_related('head', 'parent').all()
    serializer_class = DepartmentSerializer
    permission_classes = [ReadOnlyOrCanManageDepartments]
    search_fields = ['name', 'code']
    ordering_fields = ['name', 'code']
    ordering = ['name']

    @action(detail=False, methods=['get'], url_path='tree')
    def tree(self, request):
        roots = Department.objects.filter(parent__isnull=True).select_related('head')
        serializer = DepartmentTreeSerializer(roots, many=True)
        return Response(serializer.data)
